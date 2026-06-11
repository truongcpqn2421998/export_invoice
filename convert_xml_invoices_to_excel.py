import os
import glob
import xml.etree.ElementTree as ET
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

DEFAULT_FIELD_MAPPINGS = {
    'STT': 'LineNo',
    'MHHDVu': 'Code',
    'THHDVu': 'Description',
    'DVTinh': 'Unit',
    'SLuong': 'Quantity',
    'DGia': 'UnitPrice',
    'ThTien': 'Amount',
    'ThanhTien': 'Amount',
    'GiaTri': 'Amount',
    'TienThue': 'TaxAmount',
    'TSuat': 'TaxRate',
    'ThueSuat': 'TaxRate',
    'TLCKhau': 'DiscountRate',
    'STCKhau': 'DiscountAmount',
    'CKhau': 'DiscountAmount',
    'ChietKhau': 'DiscountAmount',
    'TongTien': 'TotalAmount',
    'TgTTTBSo': 'TotalAmountNumber',
    'TgTTTBChu': 'TotalAmountWords',
    'TgThue': 'TotalTaxAmount',
    'TgTT': 'TotalBeforeTax',
    'MST': 'TaxCode',
    'MaDiaDiemKD': 'BusinessLocationCode',
    'MaSoThue': 'TaxCode',
    'NguoiLap': 'IssuerName',
    'NguoiKy': 'SignerName',
    'DienGiai': 'Description',
    'AmountOC': 'AmountOC',
    'DiscountAmount': 'DiscountAmount',
}


def local_name(tag):
    if isinstance(tag, str) and '}' in tag:
        return tag.split('}', 1)[1]
    return tag


def text_of(e):
    return (e.text or '').strip() if e is not None and e.text is not None else ''


def find_elements(parent, name):
    return [el for el in parent.iter() if local_name(el.tag).lower() == name.lower()]


def extract_ttkhac(ttkhac_elem):
    # TTKhac contains TTin entries with TTruong and DLieu
    data = {}
    if ttkhac_elem is None:
        return data
    for ttin in [e for e in ttkhac_elem.iter() if local_name(e.tag) == 'TTin']:
        key = text_of(next((c for c in ttin if local_name(c.tag) == 'TTruong'), None))
        val = text_of(next((c for c in ttin if local_name(c.tag) == 'DLieu'), None))
        if key:
            data[key] = val
    return data


def normalize_item(h):
    item = {}
    # collect simple direct children by tag
    for c in list(h):
        tag = local_name(c.tag).lower()
        val = text_of(c)
        if tag in ('stt', 'lineno', 'line', 'sott'):
            item['LineNo'] = val
        elif tag in ('mhhdvu', 'mhh', 'code', 'maso'):
            item['Code'] = val
        elif tag in ('thhdvu', 'description', 'ten', 'name'):
            item['Description'] = val
        elif tag in ('dvtinh', 'unit', 'donvitinh'):
            item['Unit'] = val
        elif tag in ('sluong', 'quantity', 'qty'):
            item['Quantity'] = val
        elif tag in ('dgia', 'unitprice', 'price', 'dongia'):
            item['UnitPrice'] = val
        elif tag in ('thtien', 'amount', 'thanhtien', 'thanhtientruocthue', 'amountbeforetax'):
            item['Amount'] = val
        elif tag in ('tsuat', 'thuesuat', 'taxrate'):
            item['TaxRate'] = val
        elif tag in ('tlckhau', 'discountrate', 'discountpercent'):
            item['DiscountRate'] = val
        elif tag in ('tienthue', 'taxamount'):
            item['TaxAmount'] = val
        elif tag in ('ckhau', 'discountamount', 'discount'):
            item['DiscountAmount'] = val
        elif tag in ('tlckhau', 'discountrate', 'discountpercent'):
            item['DiscountRate'] = val
        elif tag == 'ttkhac':
            # parse inner TTKhac/TTin structures
            extra = extract_ttkhac(c)
            for k, v in extra.items():
                keyn = k.strip()
                item[keyn] = v
        else:
            # store other direct child values with their tag name
            if val:
                item[local_name(c.tag)] = val

    # also look for nested TTKhac elements anywhere inside h
    for ttk in [e for e in h.iter() if local_name(e.tag) == 'TTKhac']:
        extra = extract_ttkhac(ttk)
        for k, v in extra.items():
            item[k.strip()] = v

    # fallback: if common fields missing, try searching descendants by common names
    def get_if_missing(key, candidates):
        if item.get(key):
            return
        for cand in candidates:
            el = next((e for e in h.iter() if local_name(e.tag).lower() == cand.lower()), None)
            if el is not None and text_of(el):
                item[key] = text_of(el)
                return

    get_if_missing('LineNo', ['STT', 'LineNo'])
    get_if_missing('Code', ['MHHDVu', 'MHH'])
    get_if_missing('Description', ['THHDVu', 'Ten'])
    get_if_missing('Unit', ['DVTinh'])
    get_if_missing('Quantity', ['SLuong', 'Quantity'])
    get_if_missing('UnitPrice', ['DGia', 'UnitPrice'])
    get_if_missing('Amount', ['ThTien', 'Amount'])
    get_if_missing('TaxRate', ['TSuat', 'ThueSuat', 'TaxRate'])
    get_if_missing('TaxAmount', ['TienThue', 'TaxAmount'])
    get_if_missing('DiscountAmount', ['CKhau', 'DiscountAmount', 'Discount'])
    get_if_missing('DiscountRate', ['TLCKhau', 'DiscountRate'])

    # ensure keys exist
    for k in ('LineNo', 'Code', 'Description', 'Unit', 'Quantity', 'UnitPrice', 'Amount'):
        item.setdefault(k, '')

    return item


def find_header(root):
    # try to locate DLHDon or NDHDon container
    for e in root.iter():
        if local_name(e.tag) in ('DLHDon', 'NDHDon'):
            return e
    return root


def get_text_anywhere(elem, possible_names):
    for name in possible_names:
        el = next((e for e in elem.iter() if local_name(e.tag).lower() == name.lower()), None)
        if el is not None and text_of(el):
            return text_of(el)
    return ''


def parse_invoice(path):
    tree = ET.parse(path)
    root = tree.getroot()

    dlh = find_header(root)

    header = {}
    header['SourceFile'] = os.path.basename(path)
    header['InvoiceID'] = get_text_anywhere(dlh, ['SHDon', 'Id', 'TransactionID'])
    header['InvoiceCode'] = get_text_anywhere(dlh, ['KHDon', 'InvoiceCode', 'InvoiceSeries', 'KyHieu'])
    header['InvoiceNumber'] = get_text_anywhere(dlh, ['SHDon', 'InvoiceNo', 'InvoiceNumber'])
    header['InvoiceSeries'] = get_text_anywhere(dlh, ['MauSo', 'Series', 'InvoiceSeries'])
    header['IssueDate'] = get_text_anywhere(dlh, ['NLap', 'IssueDate', 'InvoiceDate'])
    header['InvoiceType'] = get_text_anywhere(dlh, ['THDon', 'InvoiceType'])
    header['BusinessLocationCode'] = get_text_anywhere(dlh, ['MaDiaDiemKD', 'BranchCode'])
    header['PaymentMethod'] = get_text_anywhere(dlh, ['HTThanhToan', 'PaymentMethod'])
    header['Currency'] = get_text_anywhere(dlh, ['DVTTe', 'MainCurrency', 'Currency'])
    header['ExchangeRate'] = get_text_anywhere(dlh, ['TyGia', 'ExchangeRate'])
    header['OriginalInvoiceCode'] = get_text_anywhere(dlh, ['SHDonGoc', 'OriginalInvoiceCode'])
    header['OriginalInvoiceNumber'] = get_text_anywhere(dlh, ['SOHDonGoc', 'OriginalInvoiceNumber'])
    header['AdjustmentReason'] = get_text_anywhere(dlh, ['LyDo', 'Reason'])

    # seller
    seller = next((e for e in dlh.iter() if local_name(e.tag) == 'NBan'), None)
    buyer = next((e for e in dlh.iter() if local_name(e.tag) == 'NMua'), None)
    if seller is not None:
        header['SellerName'] = get_text_anywhere(seller, ['Ten', 'Name'])
        header['SellerTaxCode'] = get_text_anywhere(seller, ['MST', 'TaxCode'])
        header['SellerAddress'] = get_text_anywhere(seller, ['DChi', 'SellerAddress'])
        header['SellerPhone'] = get_text_anywhere(seller, ['SDThoai', 'SellerPhoneNumber'])
        header['SellerEmail'] = get_text_anywhere(seller, ['Email', 'SellerEmail'])
    else:
        header['SellerName'] = header['SellerTaxCode'] = header['SellerAddress'] = header['SellerPhone'] = header['SellerEmail'] = ''

    if buyer is not None:
        header['BuyerName'] = get_text_anywhere(buyer, ['Ten', 'Name'])
        header['BuyerTaxCode'] = get_text_anywhere(buyer, ['MST', 'TaxCode'])
        header['BuyerAddress'] = get_text_anywhere(buyer, ['DChi', 'Address'])
        header['BuyerPhone'] = get_text_anywhere(buyer, ['SDThoai', 'Phone', 'BuyerPhone'])
        header['BuyerEmail'] = get_text_anywhere(buyer, ['Email', 'BuyerEmail'])
    else:
        header['BuyerName'] = header['BuyerTaxCode'] = header['BuyerAddress'] = header['BuyerPhone'] = header['BuyerEmail'] = ''

    # totals
    ttoan = next((e for e in dlh.iter() if local_name(e.tag) == 'TToan'), None)
    if ttoan is None:
        ttoan = next((e for e in root.iter() if local_name(e.tag) == 'TToan'), None)
    header['TotalBeforeTax'] = get_text_anywhere(ttoan or dlh, ['TgTT', 'TotalBeforeTax', 'TotalAmountBeforeTax'])
    header['TotalTaxAmount'] = get_text_anywhere(ttoan or dlh, ['TgThue', 'TotalTaxAmount', 'TaxAmount'])
    header['TotalAmountNumber'] = get_text_anywhere(ttoan or dlh, ['TgTTTBSo', 'TgTThue', 'TgTCThue', 'TotalAmount', 'TotalAmountNumber'])
    header['TotalAmountWords'] = get_text_anywhere(ttoan or dlh, ['TgTTTBChu', 'TotalAmountInWords', 'TotalAmountInWordsVN'])

    # items: collect all HHDVu elements anywhere
    items = []
    for h in [e for e in root.iter() if local_name(e.tag) == 'HHDVu']:
        item = normalize_item(h)
        items.append(item)

    return header, items


def convert_numeric(val, is_percentage=False):
    """Convert value to float, return 0 if fails. Handle percentage format."""
    if not val:
        return 0
    try:
        val_str = str(val).strip()
        # If it's percentage format like "8%", extract number and divide by 100
        if '%' in val_str:
            num = float(val_str.replace('%', '').replace(',', '.').strip())
            return num / 100
        else:
            num = float(val_str.replace(',', '.'))
            # If is_percentage flag, assume raw number is percentage (12.8 = 12.8%)
            if is_percentage and num > 0:
                return num / 100
            return num
    except Exception:
        return 0


def calculate_item_columns(row):
    """Calculate computed columns based on the formulas"""
    qty = convert_numeric(row.get('Quantity', 0))
    unit_price = convert_numeric(row.get('UnitPrice', 0))
    discount_rate = convert_numeric(row.get('DiscountRate', 0), is_percentage=True)
    tax_rate = convert_numeric(row.get('TaxRate', 0), is_percentage=True)
    
    # Amount before discount = Quantity * UnitPrice
    amount_before_discount = qty * unit_price
    
    # DiscountAmount = Amount before discount * DiscountRate
    discount_amount = amount_before_discount * discount_rate
    
    # Amount after discount
    amount_after_discount = amount_before_discount - discount_amount
    
    # VATAmount = (Amount after discount) * TaxRate
    vat_amount = amount_after_discount * tax_rate
    
    # Total payment = Amount after discount + VATAmount
    total_payment = amount_after_discount + vat_amount
    
    return {
        'Amount before discount': amount_before_discount,
        'DiscountRate': discount_rate,
        'DiscountAmount': discount_amount,
        'Amount after discount': amount_after_discount,
        'TaxRate': tax_rate,
        'VATAmount': vat_amount,
        'Total payment': total_payment,
    }


def convert_dir_to_excels(input_dir, output_dir=None, mapping_path=None):
    if output_dir is None:
        output_dir = input_dir
    os.makedirs(output_dir, exist_ok=True)

    xml_files = glob.glob(os.path.join(input_dir, '*.xml'))
    if not xml_files:
        print('No XML files found in', input_dir)
        return

    # default mapping is always available; optional YAML extends or overrides it
    mappings = DEFAULT_FIELD_MAPPINGS.copy()
    if mapping_path:
        try:
            import yaml
            if os.path.isfile(mapping_path):
                with open(mapping_path, 'r', encoding='utf-8') as mf:
                    data = yaml.safe_load(mf) or {}
                    file_mappings = data.get('field_mappings', {}) or {}
                    mappings.update(file_mappings)
        except Exception:
            pass

    all_headers = []
    all_items = []
    
    # Define column order for invoices_items
    header_columns = [
        'InvoiceNumber', 'InvoiceSeries', 'IssueDate', 'InvoiceType', 
        'BusinessLocationCode', 'PaymentMethod', 'Currency', 
        'OriginalInvoiceCode', 'OriginalInvoiceNumber', 'AdjustmentReason',
        'SellerName', 'SellerTaxCode', 'SellerAddress', 'SellerPhone', 'SellerEmail',
        'BuyerName', 'BuyerTaxCode', 'BuyerAddress', 'BuyerPhone', 'BuyerEmail',
        'TotalBeforeTax', 'TotalTaxAmount', 'TotalAmountNumber', 'TotalAmountWords'
    ]
    
    item_columns = [
        'Code', 'Description', 'Unit', 'Quantity', 'UnitPrice',
        'DiscountRate', 'TaxRate'
    ]
    
    computed_columns = [
        'Amount before discount', 'DiscountAmount',
        'Amount after discount', 'VATAmount', 'Total payment'
    ]
    
    all_columns = header_columns + item_columns + computed_columns

    for f in xml_files:
        header, items = parse_invoice(f)
        
        # accumulate for combined file
        for item in items:
            row = {}
            
            # Initialize all columns with empty values
            for col in all_columns:
                row[col] = ''
            
            # Fill header columns
            for col in header_columns:
                row[col] = header.get(col, '')
            
            # Fill item columns (from item data, not computed yet)
            for k, v in item.items():
                target_k = mappings.get(k, k)
                if target_k in item_columns:
                    row[target_k] = v
            
            all_items.append(row)
        
        # accumulate headers (one per invoice)
        all_headers.append(header)

    # write combined workbook
    combined_path = os.path.join(output_dir, 'all_invoices.xlsx')
    if all_headers or all_items:
        wb = Workbook()
        
        # Write invoices_summary sheet
        ws_summary = wb.active
        ws_summary.title = 'invoices_summary'
        headers_df = pd.DataFrame(all_headers)
        for r_idx, row in enumerate(headers_df.values, 1):
            for c_idx, value in enumerate(row, 1):
                ws_summary.cell(row=r_idx, column=c_idx, value=value)
        # Add header row
        for c_idx, col_name in enumerate(headers_df.columns, 1):
            ws_summary.cell(row=1, column=c_idx, value=col_name)
        
        # Write invoices_items sheet with formulas
        ws_items = wb.create_sheet('invoices_items')
        
        # Write header row
        for c_idx, col_name in enumerate(all_columns, 1):
            ws_items.cell(row=1, column=c_idx, value=col_name)
        
        # Column indices for formulas
        col_idx_map = {col: idx for idx, col in enumerate(all_columns, 1)}
        qty_col = get_column_letter(col_idx_map['Quantity'])
        unit_price_col = get_column_letter(col_idx_map['UnitPrice'])
        discount_rate_col = get_column_letter(col_idx_map['DiscountRate'])
        tax_rate_col = get_column_letter(col_idx_map['TaxRate'])
        amount_before_disc_col = get_column_letter(col_idx_map['Amount before discount'])
        discount_amt_col = get_column_letter(col_idx_map['DiscountAmount'])
        amount_after_disc_col = get_column_letter(col_idx_map['Amount after discount'])
        vat_col = get_column_letter(col_idx_map['VATAmount'])
        
        # Write data rows with formulas
        for row_idx, item in enumerate(all_items, 2):
            for c_idx, col_name in enumerate(all_columns, 1):
                if col_name == 'Amount before discount':
                    # Formula: Quantity * UnitPrice
                    ws_items.cell(row=row_idx, column=c_idx, value=f'={qty_col}{row_idx}*{unit_price_col}{row_idx}')
                elif col_name == 'DiscountAmount':
                    # Formula: Amount before discount * DiscountRate
                    ws_items.cell(row=row_idx, column=c_idx, value=f'={amount_before_disc_col}{row_idx}*{discount_rate_col}{row_idx}')
                elif col_name == 'Amount after discount':
                    # Formula: Amount before discount - DiscountAmount
                    ws_items.cell(row=row_idx, column=c_idx, value=f'={amount_before_disc_col}{row_idx}-{discount_amt_col}{row_idx}')
                elif col_name == 'VATAmount':
                    # Formula: Amount after discount * TaxRate
                    ws_items.cell(row=row_idx, column=c_idx, value=f'={amount_after_disc_col}{row_idx}*{tax_rate_col}{row_idx}')
                elif col_name == 'Total payment':
                    # Formula: Amount after discount + VATAmount
                    ws_items.cell(row=row_idx, column=c_idx, value=f'={amount_after_disc_col}{row_idx}+{vat_col}{row_idx}')
                elif col_name in ('Quantity', 'UnitPrice'):
                    # Convert to float for numeric calculations
                    val = item.get(col_name, '')
                    if val:
                        try:
                            ws_items.cell(row=row_idx, column=c_idx, value=convert_numeric(val))
                        except Exception:
                            ws_items.cell(row=row_idx, column=c_idx, value=val)
                    else:
                        ws_items.cell(row=row_idx, column=c_idx, value=0)
                elif col_name == 'DiscountRate':
                    # Format as percentage: convert to float and format
                    val = item.get(col_name, '')
                    if val:
                        try:
                            num_val = convert_numeric(val, is_percentage=True)
                            cell = ws_items.cell(row=row_idx, column=c_idx, value=num_val)
                            # Format as percentage
                            from openpyxl.styles import numbers
                            cell.number_format = '0.00%'
                        except Exception:
                            ws_items.cell(row=row_idx, column=c_idx, value=0)
                    else:
                        ws_items.cell(row=row_idx, column=c_idx, value=0)
                elif col_name == 'TaxRate':
                    # Convert to numeric format but display as percentage
                    val = item.get(col_name, '')
                    if val:
                        try:
                            num_val = convert_numeric(val, is_percentage=True)
                            cell = ws_items.cell(row=row_idx, column=c_idx, value=num_val)
                            # Format as percentage
                            from openpyxl.styles import numbers
                            cell.number_format = '0.00%'
                        except Exception:
                            ws_items.cell(row=row_idx, column=c_idx, value=val)
                    else:
                        ws_items.cell(row=row_idx, column=c_idx, value='')
                else:
                    ws_items.cell(row=row_idx, column=c_idx, value=item.get(col_name, ''))
        
        wb.save(combined_path)
        print('Wrote combined', combined_path)


if __name__ == '__main__':
    import argparse

    p = argparse.ArgumentParser(description='Convert invoice XMLs to Excel files (one Excel per XML) and a combined workbook.')
    p.add_argument('input', nargs='?', default='.', help='Input directory containing XML files')
    p.add_argument('--out', '-o', help='Output directory for Excel files (defaults to input dir)')
    p.add_argument('--mapping', '-m', help='Optional YAML mapping file to rename fields')
    args = p.parse_args()

    convert_dir_to_excels(args.input, args.out, args.mapping)
